#! /usr/bin/env python
"""
Script to generate K8 service definations by 
combining data from excel files with Jinja templates 
"""

from openpyxl import load_workbook
from jinja2 import Template

# Open up the excel file containing the data 
xwb_source_file = "port range.xlsm"
xwb = load_workbook(xwb_source_file, data_only=True)
xsheet_sr   = xwb["services"]
xsheet_pr   = xwb["ports"]

#extract columns from services sheet
svcnm = xsheet_sr['A']
lblnm = xsheet_sr['B']
slcnm = xsheet_sr['C']

#extract columns from ports sheet
prtnm = xsheet_pr['A']
prtno = xsheet_pr['B']
tptno = xsheet_pr['C']
prtcl = xsheet_pr['D']

#Import jinja2 templates
templates_dir = "./templates"
outputcfg_dir = "./outputconfigs"

#import template files
ports_template_file     = f"{templates_dir}/ports.j2"
service_template_file   = f"{templates_dir}/services.j2"

# Open up the Jinja template file (as text) and then create a Jinja Template Object for NXOS
with open(ports_template_file) as f:
    ports_template = Template(f.read(), keep_trailing_newline=True)

with open(service_template_file) as f:
    service_template = Template(f.read(), keep_trailing_newline=True)

#define j2 template as a function for access interfaces
def ports_generate (prtnm, prtno, tptno, prtcl):
    ports_config = ports_template.render(
        portname           = prtnm,
        portnumber         = prtno,
        targetportnumber   = tptno,
        protocol           = prtcl,
    )
    return (ports_config)

def service_config (svcnm, lblnm, slcnm, prtrg):
    service_config = service_template.render(
        service_name      = svcnm,
        label_name        = lblnm,
        selector_name     = slcnm,
        portrange         = prtrg,
    )
    with open(f"{outputcfg_dir}/{svcnm}" + ".yml", "w") as f:
        f.write(service_config)

#Main function to generate the configurations for all the templates
for x in range (1, xsheet_sr.max_row):
    portranges = ""

    #Loop through ports sheet 
    for y in range (1, xsheet_pr.max_row):
        portrange = ports_generate(prtnm[y].value, prtno[y].value, tptno[y].value, prtcl[y].value)
        portranges += portrange

    #save the final config
    service_config(svcnm[x].value, lblnm[x].value, slcnm[x].value, portranges)